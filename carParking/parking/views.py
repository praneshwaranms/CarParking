from django.shortcuts import render,HttpResponse,redirect
from django.urls import reverse
import pyotp
import secrets
import cv2
import pandas as pd
import serial
import pickle
import numpy as np
from .models import *
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from email.mime.text import MIMEText
from datetime import datetime
from email.mime.multipart import MIMEMultipart
import smtplib
import random
import openpyxl
from openpyxl import Workbook,load_workbook
from datetime import datetime






def generate_random_secret_key(length=16):

    random_bytes = secrets.token_bytes(length)
    base32_secret_key = pyotp.random_base32()
    return base32_secret_key

def generate_otp(secret_key):
    totp = pyotp.TOTP(secret_key)
    return totp.now()

def send_email(subject, body, to_email, smtp_server, smtp_port, sender_email, sender_password):#This is email sending function
    # Create the MIME object
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = to_email
    message['Subject'] = subject

    # Attach the body of the email
    message.attach(MIMEText(body, 'plain'))

    # Establish a connection to the SMTP server
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        try:
            # Upgrade the connection to a secure TLS connection
            server.starttls()

            # Log in to the SMTP server
            server.login(sender_email, sender_password)

            # Send the email
            server.sendmail(sender_email, to_email, message.as_string())

            print("Email sent successfully.")
        except Exception as e:
            print(f"Error sending email: {e}")

def home(request):
    # Video feed
    cap = cv2.VideoCapture(0)

    with open('CarParkPos', 'rb') as f:
        posList = pickle.load(f)

    width, height = 70, 130


    def checkParkingSpace(imgPro):
        global id_no,spaceCounter,avail_slots,notAvail_slots,camera_slots,db_slots
        camera_slots = []
        db_slots = []
        avail_slots = []
        notAvail_slots = []
         
        spaceCounter = 0
        for id_no,pos in enumerate(posList):
            x, y = pos

            imgCrop = imgPro[y:y + height, x:x + width]
            print(id_no)
            # cv2.imshow(str(x * y), imgCrop)
            count = cv2.countNonZero(imgCrop)
            camera_result = slots_booking.objects.all()
            print('availibility',camera_result[id_no].email)
            if camera_result[id_no].email != '-':
                db_slots.append(id_no)  
                notAvail_slots.append(id_no)
                print('db_slots',db_slots) 
                spaceCounter-=1
                print('not avail',notAvail_slots) 



            if count < 900:
                color = (0, 255, 0)
                thickness = 7
                spaceCounter += 1
                camera_slots.append(id_no+1)
                print('camera_slots',camera_slots)

                avail_slots = [value for value in camera_slots if value not in db_slots]

                print("avail slots",avail_slots)
                print(id_no,x,y)
            else:
                color = (0, 0, 255)
                thickness = 2
                if id_no not in notAvail_slots:
                    notAvail_slots.append(id_no+1)
                print('not',notAvail_slots)
            text = f'{id_no},{count}'

            cv2.rectangle(img, pos, (pos[0] + width, pos[1] + height), color, thickness)
            # cvzone.putTextRect(img, str(text), (x, y + height - 3), scale=1,
            #                 thickness=2, offset=0, colorR=color)

        # cvzone.putTextRect(img, f'Free: {spaceCounter}/{len(posList)}', (100, 50), scale=3,
        #                     thickness=5, offset=20, colorR=(0,200,0))
    

    success, img = cap.read()

    # img = cv2.imread(r'D:\project_carParking\IMG20240128211551.jpg')
    # img = cv2.resize(img, (800, int(img.shape[0] * (800 / img.shape[1]))))
    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    imgBlur = cv2.GaussianBlur(imgGray, (3, 3), 1)
    imgThreshold = cv2.adaptiveThreshold(imgBlur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                        cv2.THRESH_BINARY_INV, 25, 16)
    imgMedian = cv2.medianBlur(imgThreshold, 5)
    kernel = np.ones((3, 3), np.uint8)
    imgDilate = cv2.dilate(imgMedian, kernel, iterations=1)

    checkParkingSpace(imgDilate)
    cv2.imwrite("parking/static/img/frame.png",img)
    # cv2.imshow("ImageBlur", imgBlur)
    # cv2.imshow("ImageThres", imgMedian)
    result = slots_availability.objects.all()
    if result.exists():
        firstinstance = result[0]
        firstinstance.slots = avail_slots
        firstinstance.save()
    print(result[0].slots)
    # flag = request.GET.get('flag', False)
    email = request.GET.get('email', None)
    emailFromOtp = request.GET.get('emailFromOtp', None)

    time = request.GET.get('time',False)

    print('time',time)

    slot_result = slots_booking.objects.filter(email = email)
    print(email)
    print(slot_result)

    if (len(slot_result)==1) and slot_result[0].staus == 0:
        

        context = {
            'email':email
        }
        return render(request , 'otp.html', context = context)
    

    elif (len(slot_result)==1) and slot_result[0].staus == 1:
        from_time_str = slot_result[0].fromTime.strip() 
        current_time = datetime.now().time()

        # Convert 19:54 to a datetime object
        time_to_subtract = datetime.strptime(from_time_str, '%H:%M').time()

        # Create datetime objects with the same date for both current time and time to subtract
        current_datetime = datetime.combine(datetime.today(), current_time)
        subtract_datetime = datetime.combine(datetime.today(), time_to_subtract)

        # Calculate the time difference
        time_difference = current_datetime - subtract_datetime

        # Extract the minutes from the time difference
        minutes_difference = int(time_difference.total_seconds() / 60)
        

        # Combine today's date with the given time

        # Calculate the time difference in minutes


        # Calculate the time difference in minutes
        slot_result.update(duration = minutes_difference)
        print(minutes_difference)
    

        # Calculate the time difference in minutes
        context = {
            'minutes' : minutes_difference,
            'email':email

        }
        return render(request , 'countdown.html', context = context)


    # arduino_port = 'COM7'
    # ser = serial.Serial(arduino_port, 9600, timeout=1)
    if time:
        print(emailFromOtp)
        slot_result = slots_booking.objects.filter(email = emailFromOtp)
        from_time_str = slot_result[0].fromTime.strip() 
        current_time = datetime.now().time()

# Convert 19:54 to a datetime object
        time_to_subtract = datetime.strptime(from_time_str, '%H:%M').time()

        # Create datetime objects with the same date for both current time and time to subtract
        current_datetime = datetime.combine(datetime.today(), current_time)
        subtract_datetime = datetime.combine(datetime.today(), time_to_subtract)

        # Calculate the time difference
        time_difference = current_datetime - subtract_datetime

        # Extract the minutes from the time difference
        minutes_difference = int(time_difference.total_seconds() / 60)

        print(f"The difference in minutes is: {minutes_difference:.2f} minutes")
        slot_result.update(duration = minutes_difference)
        # print(time_difference_minutes)
        context = {
            'minutes' : minutes_difference,
            'email':emailFromOtp
        }
        # ser.write(b'o')
        print('this is serial')
        return render(request , 'countdown.html', context = context)
    # ser.close()€3
    paid_flag= request.GET.get('paid_flag', False) 
    login_flag = request.GET.get('login_flag', False) 
    print('paid',paid_flag,login_flag)
    email = request.GET.get('email', None)
    flag = False
    if paid_flag or login_flag:
        flag = True


    context = {
        'data':spaceCounter,
        'avail_slots' : avail_slots,
        'notAvail_slots': notAvail_slots,
        'flag': flag,
        'email': email,
    }



    return render(request,'index.html',context=context)

@csrf_exempt
def book_slot(request):
    if request.method == 'POST':

        slot_id = request.POST.get('slot_id')
        # request.session['slot_id'] = slot_id

        context = {'data': slot_id}
        print(slot_id)

        return render(request, 'bookings.html', context=context)
    
@csrf_exempt
def booked_view(request):
    if request.method == "POST":
        secret_key = generate_random_secret_key()
        otp = generate_otp(secret_key)
        slot_id = request.POST.get('slot_id')
        slot_instance = slots_booking.objects.get(slot = slot_id)
        slot_instance.email = request.POST.get('email')
        slot_instance.name = request.POST.get('name')
        slot_instance.phone_number = request.POST.get('phone_number')
        slot_instance.date = request.POST.get('date')
        slot_instance.fromTime = request.POST.get('fromTime')
        slot_instance.toTime = request.POST.get('toTime')
        slot_instance.duration = request.POST.get('duration')
        slot_instance.licensePlateNo = request.POST.get('licensePlateNo')
        slot_instance.vehicleModel = request.POST.get('vehicleModel')
        slot_instance.otp = otp
        # Save the changes to the database
        slot_instance.save()    


        subject = "Parking Allotment"
        body = f"""
        Hello {request.POST.get('name')} Your OTP is {otp}.
"""
        to_email = request.POST.get('email')
        smtp_server = "smtp.gmail.com"
        smtp_port = 587 
        sender_email = "freefireoff2020@gmail.com"
        sender_password = 'ctps wjel lklv whfg'   
        send_email(subject, body, to_email, smtp_server, smtp_port, sender_email, sender_password)
        redirect_url = reverse('home') + f'?&email={to_email}'
        return redirect(redirect_url)
        
    
def login(request):
    return render(request,'login.html')

def signup(request):
    return render(request,'signup.html')

def signup_view(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        login_instance = login_details(
            email = email,
            password = password
        )
        login_instance.save()
        print(email,password)
        return redirect('home')


def otp_view(request):
    if request.method == "POST":
        otp = request.POST.get('otp')
        email = request.POST.get('email')
        otp_result = slots_booking.objects.filter(email=email)

        if otp_result.exists() and otp_result[0].otp == otp:
            # Update 'fromTime' to the current time
            now = datetime.now()
            otp_result.update(fromTime=now.strftime(" %H:%M")) 
            otp_result.update(staus = 1)

            # Redirect to the home page with 'time' and 'email' parameters
            redirect_url = reverse('home') + f'?time=True&emailFromOtp={email}'
            return redirect(redirect_url)
        else:
            # Redirect to the OTP page with 'flag' and 'email' parameters
            context = {'flag': True, 'email': email}
            return render(request, 'otp.html', context=context)

    # Handle the case when the request method is not POST
    return render(request, 'otp.html')


def login_view(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        result = login_details.objects.filter(email = email,password=password)
        if len(result)==1:

            email = request.POST.get('email')
            redirect_url = reverse('home') + f'?login_flag=True&email={email}'
            return redirect(redirect_url)
        elif email == "admin" and password == "adminpass":
            return render(request,'download_data.html')

        context = {
            'flag':True
        }
        return render(request, 'login.html',context=context)


def car_out(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        print('payment',email)
        context = {
            'email' : email
        }
    return render(request,'payment.html',context=context)


def payment_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        result = slots_booking.objects.filter(email=email)
        data_queryset = slots_booking.objects.filter(email=email)
        excel_file_path = 'done.xlsx'

        # Save the Excel file on the server
        write_data_to_excel(data_queryset, excel_file_path)

        # Store the Excel file path in the session
        request.session['excel_file_path'] = excel_file_path

        result.update(name = '-')
        result.update(phone_number = '-')
        result.update(date = '-')
        result.update(fromTime = '-')
        result.update(toTime = '-')
        result.update(duration = '-')
        result.update(licensePlateNo = '-')
        result.update(vehicleModel = '-')
        result.update(otp = '-')
        result.update(staus = 0)
        result.update(email = '-')

        redirect_url = reverse('home') + f'?paid_flag=True&email={email}'
        return redirect(redirect_url)


def write_data_to_excel(data_queryset, excel_file_path):
    try:
        # Try to load the existing workbook
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        # If the file is new, add header row
        header = ['Email', 'Name', 'Phone Number', 'Date', 'From Time', 'To Time', 'Duration', 'License Plate No', 'Vehicle Model']
        sheet.append(header)

    # Find the last row
    last_row = sheet.max_row + 1

    # Assuming your model has fields 'email', 'name', etc.
    for index, record in enumerate(data_queryset, start=last_row):
        sheet.cell(row=index, column=1, value=record.email)
        sheet.cell(row=index, column=2, value=record.name)
        sheet.cell(row=index, column=3, value=record.phone_number)
        sheet.cell(row=index, column=4, value=record.date)
        sheet.cell(row=index, column=5, value=record.fromTime)
        sheet.cell(row=index, column=6, value=record.toTime)
        sheet.cell(row=index, column=7, value=record.duration)
        sheet.cell(row=index, column=8, value=record.licensePlateNo)
        sheet.cell(row=index, column=9, value=record.vehicleModel)
        # Repeat the above line for each field in your model

    workbook.save(excel_file_path)


def location(request):
    landmark = []
    lat = []
    lon=[]
    result = LandInformation.objects.all()
    for i in range(len(result)):
        landmark.append(result[i].landmark_name)
        lat.append(result[i].latitude)
        lon.append(result[i].longitude)

    context = {
        'name':landmark,
        'lat1':lat,
        'lon1':lon
    }

    return render(request,'location2.html',context=context)

def reg(request):

    return render(request,'registration.html')



def registration(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        aadhar_number = request.POST.get('aadhar_number')
        document_input = request.POST.get('document_input')
        landmark_name = request.POST.get('landmark_name')
        exact_location = request.POST.get('exact_location')
        lat_str, lon_str = map(float, exact_location.split(','))

        registration_instance = LandInformation(
            name = name,
            aadhar_number = aadhar_number,
            document_input= document_input,
            landmark_name = landmark_name,
            latitude = lat_str,
            longitude = lon_str,
            exact_location = exact_location
        )

        registration_instance.save()
        context = {
            'name' : name,
            'lat':lat_str,
            'lon':lon_str
        }

        # print(name,aadhar_number,document_input)
    return render(request,'success_pop.html', context=context)
    